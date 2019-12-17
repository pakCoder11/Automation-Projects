# coding=utf8
from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import socket
import os
import openpyxl #A python library to read excel file
def forcefully_wait():
    '''this function is used to wait the program'''
    count = 0
    while(count < 11):
        print("Selenium needs to be ready.. ",count,"/10")
        count += 1
        time.sleep(2)
def isFileExist(filename):
    return os.path.exists(filename)
def createFilePath(ExcelFileName):
    '''this function is used to generate the file path to be stored on desktop like C:\\Users\\PC-NAME\\Desktop\\scrapped_data.xlsx'''
    return 'C:\\Users\\'+socket.gethostname()+ '\\Desktop\\'+ExcelFileName
def assemble_excelFile(Workbook,FileName,sitename):
    '''this function is used assemble'''
    sheet = Workbook.active
    if sitename == "grubhub" or sitename == "seamless":
        try:
            sheet.cell(row=1, column=1).value = ' Title '
            sheet.cell(row=1, column=2).value = ' Description '
            sheet.cell(row=1, column=3).value = ' Ratings'
            sheet.cell(row=1, column=4).value = ' Minimum Rate'
            sheet.cell(row=1, column=5).value = ' Miles'
            sheet.cell(row=1, column=6).value = ' Estimate Time '
            sheet.cell(row=1, column=7).value = ' Restaurant Profile '
        except PermissionError:
            print("You've already open excel file - Please close the excel file in which you are looking to write data..")
    elif sitename == "delivery":
        try:
            sheet.cell(row=1, column=1).value = ' Title '
            sheet.cell(row=1, column=2).value = ' Image '
            sheet.cell(row=1, column=3).value = ' Distance '
            sheet.cell(row=1, column=4).value = ' Minimum rate'
            sheet.cell(row=1, column=5).value = ' Next Delivery time'
            sheet.cell(row=1, column=6).value = ' Restaurant Profile'
        except PermissionError:
            print("You've already open excel file - Please close the excel file in which you are looking to write data..")
    elif sitename == "yelp":
        try:
            sheet.cell(row=1, column=1).value = 'Title'
            sheet.cell(row=1, column=2).value = 'Image'
            sheet.cell(row=1, column=3).value = 'Description'
            sheet.cell(row=1, column=4).value = 'Phone number'
            sheet.cell(row=1, column=5).value = 'Address'
        except PermissionError:
            print("You've already open excel file - Please close the excel file in which you are looking to write data..")
    elif sitename == "trycaviar":
        try:
            sheet.cell(row=1, column=1).value = 'Image'
            sheet.cell(row=1, column=2).value = 'Title'
            sheet.cell(row=1, column=3).value = 'Description'
            sheet.cell(row=1, column=4).value = 'Opening Time'
        except PermissionError:
            print("You've already open excel file - Please close the excel file in which you are looking to write data..")
    Workbook.save(FileName)
    return sheet
def store_data_in_excel_file(information,sitename):
    workbook = openpyxl.Workbook()
    ExcelFileName = ""
    if sitename == "trycaviar":
        ExcelFileName = "trycaviar.xlsx"
    elif sitename == "grubhub":
        ExcelFileName = "grubub.xlsx"
    elif sitename == "seamless":
        ExcelFileName = "seamless.xlsx"
    elif sitename == "delivery":
        ExcelFileName = "delivery.xlsx"
    elif sitename == "yelp":
        ExcelFileName = "yelp.xlsx"
    ExcelFilePATH = createFilePath(ExcelFileName)
    sheet = assemble_excelFile(workbook,ExcelFileName,sitename)
    row_count = 2
    for i in range(0,len(information)):
        for j in range(0,len(information[i])):
            sheet.cell(column=j+1, row=row_count, value=information[i][j])
        row_count += 1
    workbook.save(ExcelFilePATH)
    print("Data is store in a file ",ExcelFilePATH)
def www_grubhub_com(address,ThingToEat):
    '''this method is used to crawl from www.grubhub.com'''
    '''grubhub.com and seamless.com have a same html structure it showing those restaurant which are opened at a moment'''
    data_list = []
    browser = webdriver.Chrome(executable_path=r"D:\driver\chromedriver.exe")
    '''
    browser.get("https://www.grubhub.com/")
    browser.implicitly_wait(10)
    elementID = browser.find_element_by_class_name("addressInput-textInput")
    elementID.send_keys(address)
    elementID.send_keys(Keys.ENTER)

    browser.implicitly_wait(10)
    searchButtonBar = browser.find_element_by_class_name("ghs-searchInput")
    searchButtonBar.send_keys(ThingToEat)
    searchButtonBar.send_keys(Keys.ENTER)
    browser.implicitly_wait(10)
    
    '''
    browser.get("https://www.grubhub.com/search?orderMethod=delivery&locationMode=DELIVERY&facetSet=umamiV2&pageSize=20&hideHateos=true&searchMetrics=true&queryText=chicken%20burger&latitude=38.90932846&longitude=-77.04441834&preciseLocation=true&facet=open_now%3Atrue&sortSetId=umamiv3&sponsoredSize=3&countOmittingTimes=true")
    forcefully_wait()
    div_tags = browser.find_elements_by_class_name("searchResult")
    browser.implicitly_wait(8)
    for div_tag in div_tags:
        try:
            link = WebDriverWait(div_tag, 7).until(EC.presence_of_element_located((By.CLASS_NAME, "restaurant-name"))).get_attribute("href")
            title = WebDriverWait(div_tag, 7).until(EC.presence_of_element_located((By.CLASS_NAME, "u-text-ellipsis"))).text
            description = WebDriverWait(div_tag, 7).until(EC.presence_of_element_located((By.CLASS_NAME, "u-text-secondary"))).text
            ratings = WebDriverWait(div_tag, 7).until(EC.presence_of_element_located((By.CLASS_NAME, "u-stack-x-1"))).text
            minimum_rate = WebDriverWait(div_tag, 7).until(EC.presence_of_element_located((By.CLASS_NAME, "h5"))).text
            miles = WebDriverWait(div_tag, 7).until(EC.presence_of_element_located((By.XPATH, "//*[@id='ghs-search-results-container']/div/div[2]/div/div/ghs-search-results/div[1]/div/div[3]/div/ghs-impression-tracker/div/div[2]/ghs-restaurant-card/div/div[1]/div[2]/div[2]/div[1]/ghs-restaurant-pickup-distance/span/span[1]"))).text
            est_time = WebDriverWait(div_tag, 7).until(EC.presence_of_element_located((By.CLASS_NAME, "timeEstimate"))).text
            print("Title : ", title)
            print("Restaurant link : ",link)
            print("Description : ", description)
            print("Ratings : ", ratings)
            print("Estimated time : ", est_time)
            print("Miles : ", miles)
            time.sleep(2)
            data_tuple = (title, description, minimum_rate, miles, est_time, link)
            data_list.append(data_tuple)
        except(NoSuchElementException):
            print("Data is not reach by selenium...")
            browser.implicitly_wait(4)
        time.sleep(2)
    browser.close()
    print(data_list)
    store_data_in_excel_file(data_list, "grubhub")
def www_yelp_com(address,ThingToEat):
    '''this method is used to crawl data from www.yelp.com'''
    data_list = []
    browser = webdriver.Chrome(executable_path=r"D:\driver\chromedriver.exe")
    browser.get("https://www.yelp.com/")
    browser.implicitly_wait(7)
    FindButton = browser.find_element_by_class_name("pseudo-input_field")
    FindButton.send_keys(ThingToEat)
    TypeAddressButton = browser.find_element_by_id("dropperText_Mast")
    TypeAddressButton.send_keys(address)
    TypeAddressButton.send_keys(Keys.ENTER)
    forcefully_wait()
    ul_tag = browser.find_element_by_class_name("lemon--ul__373c0__1_cxs")
    list_tags = ul_tag.find_elements_by_class_name("lemon--li__373c0__1r9wz")
    for list_tag in list_tags:
        title = ""
        image = ""
        description = ""
        address = ""
        phone_number = ""
        try:
            title = WebDriverWait(list_tag, 7).until(EC.presence_of_element_located((By.CLASS_NAME, "link-color--blue-dark__373c0__1mhJo"))).text
            image = WebDriverWait(list_tag, 7).until(EC.presence_of_element_located((By.CLASS_NAME, "photo-box-img__373c0__O0tbt"))).get_attribute("src")
            description = WebDriverWait(list_tag, 7).until(EC.presence_of_element_located((By.CLASS_NAME, "text-align--left__373c0__2pnx_"))).text
            phone_number = WebDriverWait(list_tag, 7).until(EC.presence_of_element_located((By.CLASS_NAME, "text-align--right__373c0__3ARv7"))).text
            address = WebDriverWait(list_tag, 7).until(EC.presence_of_element_located((By.CLASS_NAME, "lemon--span__373c0__3997G"))).text
        except (NoSuchElementException):
            browser.implicitly_wait(4)
            print("Data not reach by selenium...")
        except (TimeoutException):
            print("Time out")
        print("Title : ",title)
        print("Image : ",image)
        print("Description : ",description)
        print("Phone number : ",phone_number)
        print("Address : ",address)
        data_tuple = (title, image,  description, phone_number, address)
        data_list.append(data_tuple)
    browser.close()
    store_data_in_excel_file(data_list, "yelp")
def www_seamless_com(address,ThingToEat):
    '''this method is used to crawl data from www.seamless.com'''
    data_list = []
    browser = webdriver.Chrome(executable_path=r"D:\driver\chromedriver.exe")
    browser.get("https://www.seamless.com/")
    browser.implicitly_wait(10)
    elementID = browser.find_element_by_class_name("addressInput-textInput")
    elementID.send_keys(address)
    elementID.send_keys(Keys.ENTER)

    browser.implicitly_wait(10)
    searchButtonBar = browser.find_element_by_class_name("ghs-searchInput")
    searchButtonBar.send_keys(ThingToEat)
    searchButtonBar.send_keys(Keys.ENTER)

    forcefully_wait()
    div_tags = browser.find_elements_by_class_name("searchResult")
    browser.implicitly_wait(8)
    for div_tag in div_tags:
        try:
            link = WebDriverWait(div_tag, 7).until(
                EC.presence_of_element_located((By.CLASS_NAME, "restaurant-name"))).get_attribute("href")
            title = WebDriverWait(div_tag, 7).until(
                EC.presence_of_element_located((By.CLASS_NAME, "u-text-ellipsis"))).text
            description = WebDriverWait(div_tag, 7).until(
                EC.presence_of_element_located((By.CLASS_NAME, "u-text-secondary"))).text
            ratings = WebDriverWait(div_tag, 7).until(
                EC.presence_of_element_located((By.CLASS_NAME, "u-stack-x-1"))).text
            minimum_rate = WebDriverWait(div_tag, 7).until(EC.presence_of_element_located((By.CLASS_NAME, "h5"))).text
            miles = WebDriverWait(div_tag, 7).until(EC.presence_of_element_located((By.XPATH,
                                                                                    "//*[@id='ghs-search-results-container']/div/div[2]/div/div/ghs-search-results/div[1]/div/div[3]/div/ghs-impression-tracker/div/div[2]/ghs-restaurant-card/div/div[1]/div[2]/div[2]/div[1]/ghs-restaurant-pickup-distance/span/span[1]"))).text
            est_time = WebDriverWait(div_tag, 7).until(
                EC.presence_of_element_located((By.CLASS_NAME, "timeEstimate"))).text
            print("Title : ", title)
            print("Restaurant link : ", link)
            print("Description : ", description)
            print("Ratings : ", ratings)
            print("Estimated time : ", est_time)
            print("Miles : ", miles)
            time.sleep(2)
            data_tuple = (title, description, minimum_rate, miles, est_time, link)
            data_list.append(data_tuple)
        except(NoSuchElementException):
            print("Data not reach by selenium...")
            browser.implicitly_wait(4)
        time.sleep(2)
    browser.close()
    print(data_list)
    store_data_in_excel_file(data_list, "seamless")
def www_trycaviar_com(address,ThingToEat):
    '''this function is used to crawl data from www.trycaviar.com'''
    data_list = []
    browser = webdriver.Chrome(executable_path=r"D:\driver\chromedriver.exe")
    browser.get("https://www.trycaviar.com/")
    browser.implicitly_wait(7)
    elementID = browser.find_element_by_class_name("address-autocomplete_input")
    elementID.send_keys(address)
    elementID.send_keys(Keys.ENTER)
    browser.implicitly_wait(10)

    searchButton = browser.find_element_by_class_name("merchant-search_input")
    searchButton.send_keys(ThingToEat)

    forcefully_wait()
    showAllButton = browser.find_element_by_class_name("merchant-tiles-container_show-button").click()
    list_tags = browser.find_elements_by_class_name("merchant-tile")
    for list_tag in list_tags:
        image = ""
        title = ""
        description = ""
        opening_time = ""
        try:
            image = WebDriverWait(list_tag, 7).until(
                EC.presence_of_element_located((By.CLASS_NAME, "image_content"))).get_attribute("src")
            title = WebDriverWait(list_tag,7).until(
                EC.presence_of_element_located((By.CLASS_NAME,"merchant-tile_name"))).text
            description = WebDriverWait(list_tag, 7).until(
                EC.presence_of_element_located((By.CLASS_NAME, "merchant-tile_description"))).text
            opening_time = WebDriverWait(list_tag, 7).until(
                EC.presence_of_element_located((By.CLASS_NAME, "eta_opening-time"))).text
            print(image)
            print(title)
            print(description)
            print(opening_time)
        except(NoSuchElementException):
            print("Data not reach by selenium...")
            browser.implicitly_wait(4)
        data_tuple = (image,title,description,opening_time)
        data_list.append(data_tuple)
    browser.close()
    store_data_in_excel_file(data_list, "trycaviar")
    print("Data saved in excel file")
def www_delivery_com(address,ThingToEat):
    data_list = []
    browser = webdriver.Chrome(executable_path=r"D:\driver\chromedriver.exe")
    browser.get("https://www.delivery.com/")
    browser.implicitly_wait(7)
    elementID = browser.find_element_by_class_name("AutocompleteInput__input")  # AutocompleteInput__input AddressAutocomplete__Input
    elementID.send_keys(address)

    elementID = browser.find_element_by_class_name("ProcessingButton")
    elementID.submit()
    browser.implicitly_wait(7)
    areyouhungryButtonID = browser.find_element_by_class_name("SearchAutocomplete__Input")
    areyouhungryButtonID.send_keys(ThingToEat)
    areyouhungryButtonID.send_keys(Keys.ENTER)
    browser.implicitly_wait(7)
    forcefully_wait() #this is used to pause the program forefully for 20 seconds and then execute the code in order to avoid StaleElementReferenceException
    list_tags = browser.find_elements_by_class_name("SearchResults__List__Item ")
    browser.implicitly_wait(7)

    print(len(list_tags))
    for list_tag in list_tags:
        title = ""
        image = ""
        minimum_amount = ""
        mile_distance = ""
        reviews = ""
        next_delivery_time = ""
        link = ""
        try:
           title = WebDriverWait(list_tag, 7).until(
               EC.presence_of_element_located((By.CLASS_NAME, "SearchResult__TitleWrap__Title__Name"))).text
           image = WebDriverWait(list_tag, 7).until(
               EC.presence_of_element_located((By.CLASS_NAME, "SearchResult__Logo__Img"))).get_attribute("src")
           mile_distance = WebDriverWait(list_tag, 7).until(
               EC.presence_of_element_located((By.CLASS_NAME, "SearchResultDistance__Distance"))).text
           minimum_amount = WebDriverWait(list_tag, 7).until(
               EC.presence_of_element_located((By.CLASS_NAME, "SearchResultMinimum__Amount"))).text
           link = WebDriverWait(list_tag, 7).until(
               EC.presence_of_element_located((By.CLASS_NAME, "SearchResult  "))).get_attribute("href")
           reviews = WebDriverWait(list_tag, 7).until(
               EC.presence_of_element_located((By.CLASS_NAME, "SearchResultReviews--delivery"))).text
           next_delivery_time = WebDriverWait(list_tag, 7).until(
               EC.presence_of_element_located((By.CLASS_NAME, "SearchResultNextDeliveryTime__Copy__Time"))).text
        except(NoSuchElementException):
            print("Data not reach by selenium...")
            browser.implicitly_wait(4)
        except(TimeoutException):
            print("Time Out for data crawling...")
        print("Title : ", title)
        print("Product Image : ", image)
        print("Distance : ", mile_distance)
        print("Minimum amount : ", minimum_amount)
        print("Reviews : ",reviews)
        print("Next delivery time : ",next_delivery_time)
        data_tuple = (title, image, mile_distance, minimum_amount, reviews, next_delivery_time, link)
        data_list.append(data_tuple)
    browser.close()
    print(data_list)
    store_data_in_excel_file(data_list, "delivery")
    #key to fix errors is to seperately handle <ul> tags list </ul>
if __name__ == '__main__':  # Value saved here
    print("-------------------------------")
    print("----DATA CRAWLER v 1.0---------")
    print("-------------------------------")
    print("WEBSITE_KEYS \n 1 for www.delivery.com \n 4 for www.trycaviar.com\n 5 for www.yelp.com")
    website_key = input("Enter website key : ")
    address = input("Enter your street address : ")
    search_item = input("Item for search : ")
    if website_key == "1":
        print("Crawl data from www.delivery.com ")
        www_delivery_com(address,search_item)
    elif website_key == "4":
        print("Crawl data from www.trycaviar.com")
        www_trycaviar_com(address,search_item)
    elif website_key == "5":
        print("Crawl data from www.yelp.com")
        www_yelp_com(address,search_item)
    else:
        print("You entered invalid website details please enter valid one..")
    print("Program complete...")

#3 Dupont Circle NW, Washington, DC 22036
