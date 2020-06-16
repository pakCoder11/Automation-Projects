# coding=utf8
from selenium.common.exceptions import *
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
import time
from datetime import datetime
from random import randint
from openpyxl import Workbook


def generate_xpath_for_scrapping(string,integer):
    xpath = ""
    data_list = list(string)
    char = str(integer)
    data_list[113] = char
    for d in data_list:
        xpath += d
    return xpath

def google_maps_scrapping(driver,workbook,worksheet):

    ''' this function is used to perform the scrapping operations '''
    xpath_list = []

    title_list = []
    reviews_list = []
    URL_list = []
    address_list = []
    timings_list = []
    direction_details_list = []
    contacts_list = []

    divs = driver.find_element_by_xpath("/html/body/div[5]/div[2]/div[8]/div[1]/div[2]/div/div[2]/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[4]/div[1]/div")

    for i in range(0,20):
        xpath = generate_xpath_for_scrapping("/html/body/div[5]/div[2]/div[8]/div[1]/div[2]/div/div[2]/div[2]/div/div/div/div/div[2]/div/div/div[1]/div[4]/div[3]/div ",i)
        xpath_list.append(xpath) #generate the xpath list data
    
    for page in range(0,10):
        for xpath in xpath_list:
            try:
                click_on_tag = driver.find_element_by_xpath(xpath).click()
                forcefully_wait(5)

                driver.implicitly_wait(7) 
                # title = driver.find_element_by_xpath("/html/body/div[5]/div[2]/div[8]/div[1]/div[3]/div/div[2]/async-local-kp/div/div/div[1]/div/div/div/div[1]/div/div[1]/div/div[1]/div/div[1]/div/div[2]/div[1]").text
                title = driver.find_element_by_class_name("LrzXr").text

                reviews = WebDriverWait(driver, 2).until(
                EC.presence_of_element_located((By.XPATH, "/html/body/div[5]/div[2]/div[8]/div[1]/div[3]/div/div[2]/async-local-kp/div/div/div[1]/div/div/div/div[1]/div/div[1]/div/div[1]/div/div[2]/div[1]/div/div/span[2]/span/a/span "))).text

                address = WebDriverWait(driver, 2).until(
                EC.presence_of_element_located((By.XPATH, "/html/body/div[5]/div[2]/div[8]/div[1]/div[3]/div/div[2]/async-local-kp/div/div/div[1]/div/div/div/div[1]/div/div[1]/div/div[3]/div/div[2]/div/div/span[2]"))).text
            
                phone_number = WebDriverWait(driver, 2).until(
                EC.presence_of_element_located((By.XPATH, "/html/body/div[5]/div[2]/div[8]/div[1]/div[3]/div/div[2]/async-local-kp/div/div/div[1]/div/div/div/div[1]/div/div[1]/div/div[3]/div/div[4]/div/div/span[2]/span/span"))).text
            
                timings = WebDriverWait(driver, 2).until(
                EC.presence_of_element_located((By.XPATH, "/html/body/div[5]/div[2]/div[8]/div[1]/div[3]/div/div[2]/async-local-kp/div/div/div[1]/div/div/div/div[1]/div/div[1]/div/div[3]/div/div[3]/div/div/div/div/div/div[1]/div[1]/span/span/span/span/span[1]"))).text
        
                website_link_button = driver.find_element_by_xpath("/html/body/div[5]/div[2]/div[8]/div[1]/div[3]/div/div[2]/async-local-kp/div/div/div[1]/div/div/div/div[1]/div/div[1]/div/div[1]/div/div[1]/div/div[2]/div[2]/div[1]/a").click()
                forcefully_wait(2)
                
                link_href = browser.current_url
                driver.back()
            
                directions_button = WebDriverWait(driver, 2).until(
                EC.presence_of_element_located((By.XPATH, "/html/body/div[5]/div[2]/div[8]/div[1]/div[3]/div/div[2]/async-local-kp/div/div/div[1]/div/div/div/div[1]/div/div[1]/div/div[1]/div/div[1]/div/div[2]/div[2]/div[2]/a"))).click()
            
                direction_detail = driver.find_element_by_xpath("/html/body/jsl/div[3]/div[9]/div[3]/div[1]/div[2]/div/div[3]/div[1]/div[2]/div[2]/div/div/input").get_attribute("aria-label")

                print("Title " ,title)
                print("Reviews : ",reviews)
                print("Address : ",address)
                print("Phone number : ",phone_number)
                print("Timings : ",timings)
                print("Website " ,link_href)
                print("Directions ",direction_detail)            

                title_list.append(title)
                reviews_list.append(reviews)
                URL_list.append(link_href)
                direction_details_list.append(direction_detail)
                address_list.append(address)
                timings_list.append(timings)
                contacts_list.append(phone_number)

                driver.back()

            except NoSuchElementException:
                print("No elments were found...")
            except TimeoutException:
                print("Time Out")
            except ElementClickInterceptedException:
                pass
            except StaleElementReferenceException:
                pass
        print("---- SCRAPPED DATA ----")
        print(title_list,reviews_list,contacts_list,address_list,URL_list,timings_list,direction_details_list)

        write_data_in_excel_file(title_list,reviews_list,contacts_list,address_list,URL_list,timings_list,direction_details_list,workbook,worksheet)
        print("Data stores in excel file ",FileName)
        print("Data scrapped from " ,page+1)

        forcefully_wait(3)
        driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")

        driver.find_element_by_xpath('//*[@id="pnnext"]').click()
    driver.close

def forcefully_wait(times):
    '''this function is used to wait the program'''
    count = 0
    while(count < times):
        print("Selenium driver starts in seconds.. ",count+1,"/",times)
        count += 1
        time.sleep(2)

def write_data_in_excel_file(titles,reviews,contacts,addresses,URLs,timings,directions,workbook,worksheet):

    
    row_number = worksheet.max_row + 1 #get the maximum row number

    for x in range(0,len(titles)):
        worksheet.cell(column=1,row=row_number,value=titles[x])
        worksheet.cell(column=2,row=row_number,value=reviews[x])
        worksheet.cell(column=3,row=row_number,value=contacts[x])
        worksheet.cell(column=4,row=row_number,value=addresses[x])
        worksheet.cell(column=5,row=row_number,value=URLs[x])
        worksheet.cell(column=6,row=row_number,value=timings[x])
        worksheet.cell(column=7,row=row_number,value=directions[x])

        row_number += 1
        workbook.save(FileName)
if __name__ == "__main__":
    global FileName

    workbook = Workbook()

    value = randint(0, 100)
    FileName = "datafile" + str(value) + ".xlsx"
    worksheet = workbook.create_sheet("data-sheet")
    
    worksheet['A1'] = "Title"
    worksheet['B1'] = "Reviews"
    worksheet['C1'] = "Address"
    worksheet['D1'] = "Contact"
    worksheet['E1'] = "Timings"
    worksheet['F1'] = "Website Link"
    worksheet['G1'] = "Destination (Direction)"
    print("-----------------------------")
    print("Google Maps SCRAPPER")
    print("-----------------------------")

    search_data = input("Enter keyword to search e.g ' Dentist in Paris ' : ")


    browser = webdriver.Firefox(executable_path=r"C:\Users\Ahmed\bin\driver\geckodriver.exe")
    browser.get("https://www.google.com/") #open the google home page

    GoogleSearchBox = browser.find_element_by_xpath("/html/body/div/div[3]/form/div[2]/div[1]/div[1]/div/div[2]/input")
    GoogleSearchBox.send_keys(search_data)
    GoogleSearchBox.send_keys(Keys.ENTER) #press enter button

    forcefully_wait(3)

    MorePlacesLink = browser.find_element_by_class_name("i0vbXd") #find a More Places Button then 
    MorePlacesLink.click() #click on More Places Button

    google_maps_scrapping(browser,workbook,worksheet)

    browser.close
