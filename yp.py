# coding=utf8
from selenium.common.exceptions import *
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import time
import socket
import os
import openpyxl #A python library to read excel files

def assemble_excelFile(Workbook,FileName,sitename):
    '''this function is used assemble the excel file rows....'''
    sheet = Workbook.active
    if sitename == "yellowpages":
        try:
            sheet.cell(row=1, column=1).value = ' Title '
            sheet.cell(row=1, column=2).value = ' Website'
            sheet.cell(row=1, column=3).value = ' Phone Number'
            sheet.cell(row=1, column=4).value = ' Images'
            sheet.cell(row=1, column=5).value = ' Street Address '
            sheet.cell(row=1, column=6).value = ' Locality '
            sheet.cell(row=1, column=7).value = ' Description '
        except PermissionError:
            print("You've already open excel file - Please close the excel file in which you are looking to write data..")
    Workbook.save(FileName)
    return sheet
def isFileExist(filename):
    return os.path.exists(filename)
def createFilePath(ExcelFileName):
    '''this function is used to generate the file path to be stored on desktop like C:\\Users\\PC-NAME\\Desktop\\scrapped_data.xlsx'''
    return 'C:\\Users\\'+socket.gethostname()+ '\\Desktop\\'+ExcelFileName
def store_data_in_excel_file(title_list,website_list,number_list,image_list,street_addr_list,locality_list,desc_list):
    '''
    workbook = openpyxl.Workbook()
#    workbook = openpyxl.load_workbook('C:\\Users\\'+socket.gethostname()+ '\\Desktop\\'+'yellowpages.xlsx')
    #    return 'C:\\Users\\'+socket.gethostname()+ '\\Desktop\\'+ExcelFileName
    ExcelFileName = ""
    if sitename == "yellowpages":
        ExcelFileName = "yellowpages.xlsx"
    ExcelFilePATH = createFilePath(ExcelFileName)
    sheet = assemble_excelFile(workbook, ExcelFileName, sitename)
    row_number = sheet.max_row + 1
    '''
    EXCEL_FILE_PATH = 'C:\\Users\\'+socket.gethostname()+ '\\Desktop\\'+'yellowpages.xlsx'
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    sheet = workbook.active
    row_number = sheet.max_row + 1

    for x in range(0, len(title_list)):
        sheet.cell(column=1, row=row_number, value=title_list[x])
        sheet.cell(column=2, row=row_number, value=website_list[x])
        sheet.cell(column=3, row=row_number, value=number_list[x])
        sheet.cell(column=4, row=row_number, value=image_list[x])
        sheet.cell(column=5, row=row_number, value=street_addr_list[x])
        sheet.cell(column=6, row=row_number, value=locality_list[x])
        sheet.cell(column=7, row=row_number, value=desc_list[x])
        row_number += 1
    workbook.save(EXCEL_FILE_PATH)
    print("Data is store in a file ", EXCEL_FILE_PATH)
def getTitles(div_tags):
    title_list = []
    for div_tag in div_tags:
        try:
            title = WebDriverWait(div_tag, 2).until(
                EC.presence_of_element_located((By.CLASS_NAME, "business-name"))).text
            title_list.append(title)
        except (NoSuchElementException,TimeoutException):
            title_list.append("---")
    return title_list
def getDescription_list(div_tags):
    desc_list = []
    for div_tag in div_tags:
        try:
            description = WebDriverWait(div_tag, 2).until(
                EC.presence_of_element_located((By.CLASS_NAME, "body"))).text
            desc_list.append(description)
        except (NoSuchElementException,TimeoutException):
            desc_list.append("---")
    return desc_list
def getImages_list(div_tags):
    images_list = []
    for div_tag in div_tags:
        try:
            image = WebDriverWait(div_tag, 1).until(
                EC.presence_of_element_located((By.CLASS_NAME, "no-tracks"))).get_attribute("src")
            images_list.append(image)
        except (NoSuchElementException,TimeoutException):
            images_list.append("---")
    return images_list
def getPhoneNumbers_list(div_tags):
    phone_numbers_list = []
    for div_tag in div_tags:
        try:
            phone_number = WebDriverWait(div_tag, 2).until(
                EC.presence_of_element_located((By.CLASS_NAME, "phones"))).text
            phone_numbers_list.append(phone_number)
        except (NoSuchElementException,TimeoutException):
            phone_numbers_list.append("---")
    return phone_numbers_list
def getStreetAddress_list(div_tags):
    street_addresses_list = []
    for div_tag in div_tags:
        try:
            street_address = WebDriverWait(div_tag, 2).until(
                EC.presence_of_element_located((By.CLASS_NAME, "street-address"))).text
            street_addresses_list.append(street_address)
        except (NoSuchElementException,TimeoutException):
            street_addresses_list.append("---")
    return street_addresses_list
'''
def getRatings_list(div_tags):
    ratings_list = []
    for div_tag in div_tags:
        try:
            ratings = WebDriverWait(div_tag, 1).until(
                EC.presence_of_element_located((By.CLASS_NAME, "count"))).text + "ratings"
            ratings_list.append(ratings)
        except (NoSuchElementException,TimeoutException):
            ratings_list.append("---")
    return ratings_list
'''
def getWebsite_links(div_tags):
    website_list = []
    for div_tag in div_tags:
        try:
            '''
            website = WebDriverWait(div_tag, 1).until(
                EC.presence_of_element_located((By.CLASS_NAME, "track-visit-website"))).get_attribute("href")
            '''
            website = div_tag.find_element_by_class_name("track-visit-website")
            if website.is_displayed():
                website = div_tag.find_element_by_class_name("track-visit-website").get_attribute("href")
                website_list.append(website)
        except (NoSuchElementException,TimeoutException):
            website_list.append("---")
    return website_list
def get_locality_list(div_tags):
    locality_list = []
    for div_tag in div_tags:
        try:
            locality = WebDriverWait(div_tag, 2).until(
                EC.presence_of_element_located((By.CLASS_NAME, "locality"))).text
            locality_list.append(locality)
        except (NoSuchElementException,TimeoutException):
            locality_list.append("---")
    return locality_list
'''
def getDivtags(driver):
    url_list = ["www.yellowpages.com"]
    while True:
        if url_list[len(url_list)-1] == driver.current_url:
            break
        else:
            url_list.append(driver.current_url)
            print("Page lookup ",driver.current_url)
            NextButton = driver.find_element_by_class_name("next").click()
            print("Last url visit" , url_list[len(url_list)-1])
            forcefully_wait(4)
    return url_list
'''
def generateLinks(driver):
    url_links = []
    for i in range(1,100):
        page = driver.current_url+'&page='+str(i)
        url_links.append(page)
    return url_links
def forcefully_wait(times):
    '''this function is used to wait the program'''
    count = 0
    while(count < times):
        print("Selenium needs to be ready.. ",count+1,"/",times)
        count += 1
        time.sleep(2)
def www_yellow_pages_com(ThingToSearch,address): #this function is used to crawl data from Yellowpages.com
    workbook = openpyxl.Workbook()

    browser = webdriver.Chrome(ChromeDriverManager().install())
    print("Web Driver Install")
    browser.get("https://www.yellowpages.com/")
    browser.implicitly_wait(7) #wait for 7 seconds until DOM elements in the page are getting ready.

    SearchThingsBar = browser.find_element_by_name("search_terms")
    SearchThingsBar.send_keys(ThingToSearch)
    SearchThingsBar.send_keys(Keys.TAB)
    AddressInputBar = browser.find_element_by_name("geo_location_terms")
    AddressInputBar.send_keys(address)
    AddressInputBar.send_keys(Keys.ENTER)
    forcefully_wait(3)
    URLs = generateLinks(browser)
    print(URLs)
    for url in URLs:
        if browser.current_url == url:
            browser.close()
            break
        browser.get(url)
        forcefully_wait(7)
        os.system("cls")
        print("Processing.... It will take 4-5 minutes to crawl data from ",url)
        
        div_tags = browser.find_elements_by_class_name("result")
        title_list = getTitles(div_tags)
        number_list = getPhoneNumbers_list(div_tags)
        street_addr_list = getStreetAddress_list(div_tags)
        desc_list = getDescription_list(div_tags)
        locality_list = get_locality_list(div_tags)
        image_list = getImages_list(div_tags)
        website_list = getWebsite_links(div_tags)

        print("Data crawled from ",url)
        print(title_list)
        print(number_list)
        print(street_addr_list)
        print(desc_list)
        print(locality_list)
        print(image_list)
        print(website_list)
        try:
            store_data_in_excel_file(title_list, website_list, number_list, image_list, street_addr_list, locality_list,
                                 desc_list)
            excel_file = 'C:\\Users\\'+socket.gethostname()+ '\\Desktop\\'+'yellowpages.xlsx'
            print("Data stored in excel file ",excel_file)
        except PermissionError:
            print("Sorry! Data failed to write in a excel becase file is already open. Please close ",excel_file)
    browser.close()
if __name__ == '__main__':
    print("-------------------------------")
    print("----YP Lead Gen.---------")
    print("-------------------------------")
    website_key = input("Enter keyword to search e.g ' Real Estate ' : ")
    location = input("Enter your location e.g ' New York, NY' : ")
    www_yellow_pages_com(website_key,location)
    print("Thanks . SAY NO TO #islamophobia")